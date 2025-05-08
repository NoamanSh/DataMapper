import pandas as pd
import xml.etree.ElementTree as ET
import os
from openpyxl import load_workbook

def extract_data_from_one_xml(xml_file_path, mappings_for_this_xml):
    """
    Extracts data from a single XML file based on provided mappings.
    The number of rows generated is determined by the XML data.

    Args:
        xml_file_path (str): Path to the XML file.
        mappings_for_this_xml (dict): Dictionary mapping sheet names to column mappings
            Example: {'Sheet1': {'Column1': 'xml/path1', 'Column2': 'xml/path2'}}

    Returns:
        dict: Dictionary mapping sheet names to pandas DataFrames containing extracted data.
    """
    # print(f"DEBUG: Extracting data from XML: {xml_file_path}")
    # print(f"DEBUG: Starting apply_mapping_and_fill. Output will be: {output_path}") # Keep for debugging if needed
    
    # Parse XML file
    tree = ET.parse(xml_file_path)
    root = tree.getroot()
    
    # This will hold DataFrames for each sheet, ready to be written
    # Key: sheet_name, Value: pd.DataFrame
    data_to_write_to_excel = {}

    for sheet_name, sheet_mappings in mappings_for_this_xml.items():
        # print(f"\nDEBUG: Processing sheet: {sheet_name}") # Keep for debugging
        # For this sheet, collect all data columns from XML
        # Key: excel_col_name (target header), Value: list of extracted values
        raw_sheet_data = {}
        max_len = 0

        for excel_col, xml_tag_path in sheet_mappings.items():
            values = []
            # print(f"  DEBUG: Mapping Excel column '{excel_col}' to XML path '{xml_tag_path}'") # Keep for debugging
            
            path_for_findall = ""
            is_attribute_extraction = '@' in xml_tag_path
            attr_name = None

            # Handle attributes (tags with @)
            if is_attribute_extraction:
                base_xml_path, attr_name = xml_tag_path.rsplit('/@', 1)
                if base_xml_path == root.tag: # Attribute is on the root element
                    path_for_findall = "." 
                elif base_xml_path.startswith(root.tag + '/'): # Attribute is on a child/descendant
                    path_for_findall = base_xml_path[len(root.tag) + 1:] # Path relative to root
                else:
                    # This case implies base_xml_path is already relative or a direct child tag name
                    # This might occur if utils.extract_xml_tags changes its output format
                    path_for_findall = base_xml_path 
                # print(f"    DEBUG: Attribute path. Element search path for findall: '{path_for_findall}', Attribute: '{attr_name}'")
            else:
                # Regular element text extraction
                if xml_tag_path == root.tag: 
                     path_for_findall = "."
                elif xml_tag_path.startswith(root.tag + '/'):
                    path_for_findall = xml_tag_path[len(root.tag) + 1:]
                else:
                    path_for_findall = xml_tag_path 
                # print(f"    DEBUG: Element text path. Search path for findall: '{path_for_findall}'")
            
            find_elements = root.findall(path_for_findall)
            # print(f"    DEBUG: Found {len(find_elements)} elements for path '{path_for_findall}'.")

            if is_attribute_extraction:
                for elem in find_elements:
                    values.append(elem.get(attr_name)) # Use .get() for safety, defaults to None
            else:
                for elem in find_elements:
                    values.append(elem.text if elem.text is not None else "") # Ensure empty string for None text
            
            # print(f"    DEBUG: Extracted values for '{excel_col}': {str(values[:5])[:100]}... (first 5, truncated)")
            raw_sheet_data[excel_col] = values
            if len(values) > max_len:
                max_len = len(values)
        # print(f"  DEBUG: Max data length for sheet '{sheet_name}': {max_len}")
        # Normalize lengths for all columns in this sheet
        aligned_sheet_data = {}
        for col_name, val_list in raw_sheet_data.items():
            # Pad with None if shorter, truncate if longer (though truncation shouldn't happen with max_len logic)
            aligned_sheet_data[col_name] = (val_list + [None] * (max_len - len(val_list)))[:max_len]
        if aligned_sheet_data: # only create DataFrame if there's data
            data_to_write_to_excel[sheet_name] = pd.DataFrame(aligned_sheet_data)
        else: # No data for any mapped column in this sheet
            data_to_write_to_excel[sheet_name] = pd.DataFrame() # Empty DataFrame
    
    return data_to_write_to_excel

def write_dataframes_to_excel(excel_template_file, data_frames_by_sheet, output_path):
    """
    Writes data from pandas DataFrames to an Excel template file.

    Args:
        excel_template_file (str): Path to the Excel template file.
        data_frames_by_sheet (dict): Dictionary mapping sheet names to pandas DataFrames.
        output_path (str): Path where the filled Excel file will be saved.

    Returns:
        str: Path to the output Excel file.
    """
    # print(f"DEBUG: Writing DataFrames to Excel template: {excel_template_file}, Output: {output_path}")
    # Load the template workbook
    workbook = load_workbook(excel_template_file)

    for sheet_name_in_template in workbook.sheetnames:
        if sheet_name_in_template in data_frames_by_sheet:
            # print(f"\nDEBUG: Writing to template sheet: '{sheet_name_in_template}'")
            worksheet = workbook[sheet_name_in_template]
            df_for_sheet = data_frames_by_sheet[sheet_name_in_template]

            if df_for_sheet.empty:
                # print(f"  DEBUG: DataFrame for sheet '{sheet_name_in_template}' is empty. Skipping.")
                continue # Skip writing if no data was extracted for this sheet's mappings

            # Get header row from template to map column names to indices
            header_row_values = [cell.value for cell in worksheet[1]] # First row (header)
            header_index_map = {header: col_idx for col_idx, header in enumerate(header_row_values, start=1) if header is not None}
            # print(f"  DEBUG: Template headers for '{sheet_name_in_template}': {header_index_map}")

            for excel_col_header in df_for_sheet.columns: # These are the target Excel headers from UI mapping
                # print(f"    DEBUG: Attempting to write data for Excel column: '{excel_col_header}'")
                if excel_col_header in header_index_map:
                    col_idx = header_index_map[excel_col_header]
                    # print(f"      DEBUG: Matched template header '{excel_col_header}' at column index {col_idx}.")
                    for row_idx_df, value in enumerate(df_for_sheet[excel_col_header]):
                        # row_idx_excel starts from 2 (data rows, 1-based index for openpyxl)
                        worksheet.cell(row=row_idx_df + 2, column=col_idx, value=value)
                else:
                    pass # print(f"      DEBUG: WARNING - Column '{excel_col_header}' (from UI mapping/DataFrame) not found in template sheet '{sheet_name_in_template}' headers. This column's data will NOT be written.")
        else:
            pass # print(f"\nDEBUG: Sheet '{sheet_name_in_template}' from template has no mappings/data from XML. Skipping.")
    workbook.save(output_path)
    # print(f"DEBUG: Workbook saved to {output_path}")
    return output_path